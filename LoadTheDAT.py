# -*- coding:utf-8 -*-
from TimeToName import time_name
import openpyxl
import time
name = time_name(2017, 12, 16)
CZ = []         # 厂站
line2 = []
WLL = []  # 物理量（即电压、电流、有功等）
ZHI = []  # 值
Path = r'D:\Python Study\利用Python读取EMS数据\01 20171217EMS数据'

start = time.clock()
range_end = 50
for i in range(0, range_end):
    print(i)
    file = Path + '\EMS_15M_' + name[i] + '.dat'
    EMS1 = open(file, 'r')
    line = EMS1.read()
    line = line.splitlines()
    line1 = line[2:-1]      # 去掉前两行和最后一行
    line2 = []

    if i == 0:
        l = len(line1)
        for x in line1:
            line2.append(x.split(','))      # append方法表示从列表最后新增元素
            CZ.append([line2[-1][1]])         # 厂站位于第二列，Python计数从0开始
            WLL.append([line2[-1][2]])        # 物理量位于第三列
            ZHI.append(list(map(float, [line2[-1][-1]])))       # 值在最后一列
    else:
        j = 0
        for x in line1:
            line2.append(x.split(','))  # append方法表示从列表最后新增元素
            CZ[j].append(line2[-1][1])  # 厂站位于第二列，Python计数从0开始
            WLL[j].append(line2[-1][2])  # 物理量位于第三列
            ZHI[j].append(float(line2[-1][-1]))  # 值在最后一列
            j = j + 1
    EMS1.close()


wb = openpyxl.Workbook()
ws = wb.active
ws.title = '厂站'
ws1 = wb.create_sheet('物理量')
ws2 = wb.create_sheet('值')

for i in range(0, len(CZ)):
    # ws.append(CZ[i])
    # ws1.append(WLL[i])
    ws2.append(ZHI[i])

wb.save('try.xlsx')

elapsed = (time.clock() - start)

print("Time used:", elapsed)
# print(CZ)
# print(WLL)
# print(ZHI)
# print(line2[-1])
